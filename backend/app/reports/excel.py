"""openpyxl report generators: the daily live-trading monitoring sheet and
the backtest summary workbook."""
from __future__ import annotations

import datetime as dt
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.backtest.stats import BacktestStats
from app.config import settings

HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)

DAILY_COLUMNS = [
    ("Date", "trade_date"),
    ("Stock/Index", "symbol_label"),
    ("Entry Time", "entry_time"),
    ("Entry Price", "entry_price"),
    ("Exit Time", "exit_time"),
    ("Exit Price", "exit_price"),
    ("Stop Loss", "stop_loss"),
    ("Take Profit", "take_profit"),
    ("Direction", "direction"),
    ("Entry Type", "entry_type"),
    ("Reason for Entry", "entry_reason"),
    ("Reason for Exit", "exit_reason"),
    ("P/L (Points)", "pnl_points"),
    ("P/L (Amount)", "pnl_amount"),
    ("R:R Achieved", "rr_achieved"),
    ("Screenshot/Chart Ref", "snapshot_path"),
    ("Trade Status", "status"),
]


def _style_header(ws, columns: list[str]) -> None:
    for col_idx, name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"


def _autofit(ws, n_cols: int, width: int = 18) -> None:
    for i in range(1, n_cols + 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def _fmt(value):
    if isinstance(value, (dt.datetime, dt.date)):
        return value.strftime("%Y-%m-%d %H:%M:%S") if isinstance(value, dt.datetime) else value.isoformat()
    if isinstance(value, float):
        return round(value, 2)
    return value


def write_daily_sheet(trades: list[dict], report_date: dt.date, embed_charts: bool = True) -> Path:
    """One row per trade for the given date. Overwrites any existing report
    for that date so re-runs (e.g. the 17:00 finalization job) stay
    idempotent."""
    wb = Workbook()
    ws = wb.active
    ws.title = report_date.isoformat()

    headers = [h for h, _ in DAILY_COLUMNS]
    _style_header(ws, headers)
    _autofit(ws, len(headers))
    ws.column_dimensions[get_column_letter(len(headers))].width = 24  # snapshot column wider

    row_idx = 2
    for t in trades:
        for col_idx, (_, key) in enumerate(DAILY_COLUMNS, start=1):
            if key == "snapshot_path":
                continue
            ws.cell(row=row_idx, column=col_idx, value=_fmt(t.get(key)))

        snap = t.get("snapshot_path")
        if embed_charts and snap and Path(snap).exists():
            try:
                img = XLImage(snap)
                img.width, img.height = 220, 115
                col_letter = get_column_letter(len(DAILY_COLUMNS))
                ws.add_image(img, f"{col_letter}{row_idx}")
                ws.row_dimensions[row_idx].height = 90
            except Exception:
                ws.cell(row=row_idx, column=len(DAILY_COLUMNS), value=snap)
        elif snap:
            ws.cell(row=row_idx, column=len(DAILY_COLUMNS), value=snap)

        row_idx += 1

    out_path = settings.reports_dir / f"daily_{report_date.isoformat()}.xlsx"
    wb.save(out_path)
    return out_path


def write_backtest_workbook(
    trades: list[dict], stats: BacktestStats, start_date: dt.date, end_date: dt.date, run_id: int
) -> Path:
    wb = Workbook()

    # --- Summary sheet -------------------------------------------------------
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = f"Backtest Summary: {start_date.isoformat()} to {end_date.isoformat()}"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:B1")

    summary_rows = [
        ("Total Trades", stats.total_trades),
        ("Winning Trades", stats.winning_trades),
        ("Losing Trades", stats.losing_trades),
        ("Win Rate (%)", round(stats.win_rate_pct, 2)),
        ("Loss Rate (%)", round(stats.loss_rate_pct, 2)),
        ("Total Profit (Points)", round(stats.total_profit_points, 2)),
        ("Total Loss (Points)", round(stats.total_loss_points, 2)),
        ("Net Profit (Points)", round(stats.net_profit_points, 2)),
        ("Net Profit (Amount)", round(stats.net_profit_amount, 2)),
        ("Max Winning Streak", stats.max_winning_streak),
        ("Max Losing Streak", stats.max_losing_streak),
        ("Avg Profit per Trade (Points)", round(stats.avg_profit_per_trade, 2)),
        ("Avg Loss per Trade (Points)", round(stats.avg_loss_per_trade, 2)),
        ("Max Drawdown (Points)", round(stats.max_drawdown_points, 2)),
    ]
    for i, (label, value) in enumerate(summary_rows, start=3):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws.cell(row=i, column=2, value=value)
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 18

    reduced_days = sum(1 for t in trades if t.get("reduced_resolution"))
    if reduced_days:
        note_row = len(summary_rows) + 5
        ws.cell(
            row=note_row,
            column=1,
            value=(
                f"Note: {reduced_days} day(s) in this range used 5-minute candles instead of 1-minute "
                "for structure/entry analysis, because Yahoo Finance only serves 1-minute intraday "
                "data for the trailing ~30 days. See the 'Reduced Resolution' column in the Trade Log."
            ),
        ).font = Font(italic=True, color="B45309")
        ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=6)
        ws.row_dimensions[note_row].height = 30

    # --- Trade Log sheet -------------------------------------------------------
    ws2 = wb.create_sheet("Trade Log")
    log_columns = [h for h, _ in DAILY_COLUMNS] + ["Reduced Resolution"]
    _style_header(ws2, log_columns)
    _autofit(ws2, len(log_columns))

    row_idx = 2
    for t in trades:
        for col_idx, (_, key) in enumerate(DAILY_COLUMNS, start=1):
            ws2.cell(row=row_idx, column=col_idx, value=_fmt(t.get(key)))
        ws2.cell(row=row_idx, column=len(log_columns), value=bool(t.get("reduced_resolution")))
        row_idx += 1

    # --- Monthly Performance sheet ----------------------------------------------
    ws3 = wb.create_sheet("Monthly Performance")
    monthly_headers = ["Month", "Trades", "Wins", "Losses", "Win Rate (%)", "Net P/L (Points)", "Net P/L (Amount)"]
    _style_header(ws3, monthly_headers)
    _autofit(ws3, len(monthly_headers))
    row_idx = 2
    for month, m in stats.monthly.items():
        ws3.cell(row=row_idx, column=1, value=month)
        ws3.cell(row=row_idx, column=2, value=m["trades"])
        ws3.cell(row=row_idx, column=3, value=m["wins"])
        ws3.cell(row=row_idx, column=4, value=m["losses"])
        ws3.cell(row=row_idx, column=5, value=round(m["win_rate_pct"], 2))
        ws3.cell(row=row_idx, column=6, value=round(m["net_points"], 2))
        ws3.cell(row=row_idx, column=7, value=round(m["net_amount"], 2))
        row_idx += 1

    out_path = settings.reports_dir / f"backtest_{run_id}_{start_date.isoformat()}_{end_date.isoformat()}.xlsx"
    wb.save(out_path)
    return out_path
